"""
Servidor Flask – Sistema de Liquidación de Nómina
"""

import io
import json
import os
import re
import sqlite3
import unicodedata
import zipfile
from difflib import get_close_matches
from datetime import datetime, date
from pathlib import Path

from flask import Flask, jsonify, request, send_file, render_template
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from engine import TURNO_CATALOGO, calcular_nomina

app = Flask(__name__)

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"

# ─────────────────────────────────────────────────────────────
# Base de datos SQLite para persistencia de configuración
# ─────────────────────────────────────────────────────────────

# En producción (Render) se puede montar un disco en /data.
# En local sigue usando la carpeta data/ del proyecto.
DB_PATH = Path(os.environ.get("DB_PATH", str(DATA_DIR / "nomina.db")))


def _get_db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def _init_db():
    with _get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS config (
                key   TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
        """)
        conn.commit()


def _db_get(key):
    with _get_db() as conn:
        row = conn.execute("SELECT value FROM config WHERE key = ?", (key,)).fetchone()
        return json.loads(row["value"]) if row else None


def _db_set(key, value):
    with _get_db() as conn:
        conn.execute(
            "INSERT INTO config(key, value) VALUES(?, ?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, json.dumps(value, ensure_ascii=False)),
        )
        conn.commit()


# ─────────────────────────────────────────────────────────────
# Utilidades
# ─────────────────────────────────────────────────────────────

def _load_festivos():
    fp = DATA_DIR / "festivos.json"
    if fp.exists():
        return json.loads(fp.read_text(encoding="utf-8"))
    return []


def _load_config():
    """Configuración persistente desde SQLite."""
    return _db_get("user_config") or {}


def _save_config(cfg):
    _db_set("user_config", cfg)


def _load_current():
    data = _load_defaults()
    cfg = _load_config()
    if cfg:
        if isinstance(cfg.get("params"), dict):
            cfg_params = dict(cfg["params"])
            # Fase 5: anio y mes siempre se toman del sistema, ignorar lo persistido
            cfg_params.pop("anio", None)
            cfg_params.pop("mes", None)
            data["params"].update(cfg_params)
        if isinstance(cfg.get("reglas"), dict):
            data["reglas"].update(cfg["reglas"])
        if isinstance(cfg.get("festivos"), list):
            data["festivos"] = cfg["festivos"]
    return data


def _load_defaults():
    hoy = date.today()
    return {
        "params": {
            "anio": hoy.year,
            "mes": hoy.month,
            "dias_ciclo": 21,
            "horas_objetivo": 132,
            "horas_turno": 12.583333,
            "smmlv": 1423500,
            "auxilio_transporte_mensual": 200000,
            "inicio_noc_h": 19,
            "fin_noc_h": 6,
            "almuerzo_offset_h": 6,
            "almuerzo_duracion_min": 25,
        },
        "reglas": {
            "REC_DIURNO": 0.0,
            "REC_NOCTURNO": 0.35,
            "REC_DOM_FEST_DIURNO": 0.80,
            "REC_DOM_FEST_NOCTURNO": 1.15,
            "EXT_DIURNA": 1.25,
            "EXT_NOCTURNA": 1.75,
            "EXT_FEST_DIURNA": 2.05,
            "EXT_FEST_NOCTURNA": 2.55,
        },
        "festivos": _load_festivos(),
        "empleados": [],
    }


def _fecha_dmy(iso):
    s = str(iso or "")[:10]
    p = s.split("-")
    return f"{p[2]}/{p[1]}/{p[0]}" if len(p) == 3 else s


def _parse_fecha(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return ""
    s = s.split(" ")[0]
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%y", "%d-%m-%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return s[:10]


TURNOS_VALIDOS = set(TURNO_CATALOGO.keys())


def _normalizar_codigo_turno(val):
    if val is None:
        return "", None
    raw = str(val).strip()
    if not raw:
        return "", None
    norm = "".join(raw.upper().split())
    if norm in TURNOS_VALIDOS:
        issue = None
        if norm != raw:
            issue = {
                "type": "warning",
                "message": f"Se normalizó '{raw}' a '{norm}'.",
                "suggestion": norm,
            }
        return norm, issue
    suggestion = get_close_matches(norm, list(TURNOS_VALIDOS), n=1, cutoff=0.75)
    issue = {
        "type": "error",
        "message": f"Turno '{raw}' no reconocido.",
    }
    if suggestion:
        issue["suggestion"] = suggestion[0]
        issue["message"] += f" Sugerencia: '{suggestion[0]}'."
    return norm, issue


def _validar_turnos_empleados(empleados):
    cleaned = []
    warnings = []
    errors = []
    empleados_afectados = set()
    total_celdas = 0
    normalizadas = 0

    for emp in empleados:
        emp_clean = dict(emp)
        turnos = {}
        for d, val in (emp.get("turnos") or {}).items():
            total_celdas += 1
            norm, issue = _normalizar_codigo_turno(val)
            if norm:
                turnos[str(d)] = norm
            if issue:
                base = {
                    "employee_id": emp.get("id"),
                    "employee_name": emp.get("nombre", ""),
                    "row": emp.get("_source_row"),
                    "day": int(d) if str(d).isdigit() else d,
                    "value": val,
                    "normalized": norm,
                }
                base.update(issue)
                if issue["type"] == "warning":
                    warnings.append(base)
                    normalizadas += 1
                else:
                    errors.append(base)
                empleados_afectados.add(emp.get("id"))
        emp_clean["turnos"] = turnos
        cleaned.append(emp_clean)

    validation = {
        "ok": len(errors) == 0,
        "warnings": warnings,
        "errors": errors,
        "summary": {
            "empleados": len(empleados),
            "empleados_afectados": len([x for x in empleados_afectados if x is not None]),
            "celdas_turno": total_celdas,
            "advertencias": len(warnings),
            "errores": len(errors),
            "normalizadas": normalizadas,
        },
    }
    return cleaned, validation


def _excel_to_data(wb):
    data = _load_current()
    sheet_name = "PLANTILLA_CARGA" if "PLANTILLA_CARGA" in wb.sheetnames else "MATRIZ_CARGA"
    turno_col_offset = 7 if sheet_name == "PLANTILLA_CARGA" else 9
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = None
        empleados = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row[0] == "ID":
                header_row = i
                continue
            if header_row and row[0] and str(row[0]).isdigit() or (header_row and isinstance(row[0], (int, float)) and row[0]):
                emp_id = row[0]
                if not isinstance(emp_id, (int, float)):
                    continue
                nombre = str(row[1] or "").strip()
                salario = float(row[2] or 0)
                fic = _parse_fecha(row[3])
                saldo = float(row[4] or 0)
                dias_trab = None
                aux_manual = None
                if sheet_name == "PLANTILLA_CARGA":
                    dias_trab = row[5] if len(row) > 5 and row[5] else None
                    aux_manual = row[6] if len(row) > 6 and row[6] else None
                turnos = {}
                for d in range(1, 32):
                    val = row[turno_col_offset + d - 1] if len(row) > turno_col_offset + d - 1 else None
                    if val:
                        turnos[str(d)] = str(val).strip()
                emp_dict = {
                    "id": int(emp_id),
                    "nombre": nombre,
                    "salario_mensual": salario,
                    "fecha_inicio_ciclo": fic,
                    "saldo_inicial_horas": saldo,
                    "turnos": turnos,
                    "_source_row": i,
                }
                if dias_trab is not None:
                    try: emp_dict["dias_trabajados_manual"] = int(dias_trab)
                    except: pass
                if aux_manual is not None:
                    try: emp_dict["auxilio_transporte_manual"] = float(aux_manual)
                    except: pass
                empleados.append(emp_dict)
        data["empleados"], data["validation"] = _validar_turnos_empleados(empleados)
    return data


def _generar_excel(resultados, params, reglas=None):
    reglas = reglas or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LIQUIDACION"

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT  = Alignment(horizontal="right",  vertical="center")
    thin   = Side(style="thin", color="BFBFBF")
    BDR    = Border(left=thin, right=thin, top=thin, bottom=thin)

    C_TITLE = "1F4E79"; C_REC = "1d6fa0"; C_EXT = "2585c0"; C_MON = "1F4E79"
    C_TOT = "c6efce"; C_ALT = "f7fafe"

    def hdr(fg, sz=9, color="FFFFFF"):
        return Font(bold=True, color=color, size=sz), PatternFill("solid", fgColor=fg)

    def pc(k):
        return round(float(reglas.get(k, 0)) * 100)

    mes_nombre = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
                  "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"][int(params["mes"])]

    cols = [
        ("ID", C_TITLE), ("EMPLEADO", C_TITLE), ("SALARIO BASE", C_TITLE),
        (f"REC. NOCTURNO {pc('REC_NOCTURNO')}%", C_REC),
        (f"REC. DOM/FEST {pc('REC_DOM_FEST_DIURNO')}%", C_REC),
        (f"REC. DOM/FEST NOCT {pc('REC_DOM_FEST_NOCTURNO')}%", C_REC),
        ("TOTAL RECARGOS", C_REC),
        (f"H.E DIURNA {pc('EXT_DIURNA')}%", C_EXT),
        (f"H.E NOCTURNA {pc('EXT_NOCTURNA')}%", C_EXT),
        (f"H.E FEST DIURNA {pc('EXT_FEST_DIURNA')}%", C_EXT),
        (f"H.E FEST NOCT {pc('EXT_FEST_NOCTURNA')}%", C_EXT),
        ("TOTAL H.E", C_EXT),
        ("AUXILIO TRANSP.", C_MON),
        ("TOTAL A PAGAR", C_MON),
    ]
    ncol = len(cols)
    last_col = get_column_letter(ncol)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"LIQUIDACION DE NOMINA – {mes_nombre} {params['anio']} · PERSONAL DE SEGURIDAD"
    f, fill = hdr(C_TITLE, sz=12); ws["A1"].font = f; ws["A1"].fill = fill; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (f"Ciclo: {params['dias_ciclo']} días  |  Horas objetivo: {params['horas_objetivo']}h  |  "
                f"IVH = Salario ÷ {params.get('horas_mes', 220)}  |  "
                f"Nocturno: {params.get('inicio_noc_h', 19)}:00–{params.get('fin_noc_h', 6)}:00")
    f2, fill2 = hdr("2E75B6", sz=8); ws["A2"].font = f2; ws["A2"].fill = fill2; ws["A2"].alignment = CENTER
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 30
    for c, (lbl, fg) in enumerate(cols, 1):
        cell = ws.cell(row=3, column=c, value=lbl)
        f, fill = hdr(fg); cell.font = f; cell.fill = fill; cell.alignment = CENTER; cell.border = BDR

    FMT_NUM = "#,##0"
    FILL_ALT = PatternFill("solid", fgColor=C_ALT)
    tot = [0] * ncol
    start_row = 4

    for i, r in enumerate(resultados, start=start_row):
        v = r.get("val", {})
        row_data = [
            r["id"], r["nombre"], round(r["salario_mensual"]),
            v.get("rec_noct", 0), v.get("rec_fest_d", 0), v.get("rec_fest_n", 0), r["valor_recargo"],
            v.get("ext_diurna", 0), v.get("ext_noct", 0), v.get("ext_fest_d", 0), v.get("ext_fest_n", 0), r["valor_extra"],
            r["auxilio_transporte"], r["total_pagar"],
        ]
        is_even = (i % 2 == 0)
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BDR
            if is_even and col > 1:
                cell.fill = FILL_ALT
            if col == 1:
                cell.alignment = CENTER
            elif col == 2:
                cell.alignment = Alignment(vertical="center")
            else:
                cell.alignment = RIGHT; cell.number_format = FMT_NUM
                if col in (7, 12):
                    cell.font = Font(bold=True, size=9)
                if col == ncol:
                    cell.font = Font(bold=True, color="C00000", size=9)
            if col >= 3 and isinstance(val, (int, float)):
                tot[col - 1] += val

    tot_row = start_row + len(resultados)
    ws.row_dimensions[tot_row].height = 18
    ws.merge_cells(f"A{tot_row}:B{tot_row}")
    cell = ws.cell(row=tot_row, column=1, value=f"TOTALES ({len(resultados)} empleados)")
    cell.font = Font(bold=True, size=9); cell.fill = PatternFill("solid", fgColor=C_TOT)
    cell.alignment = CENTER; cell.border = BDR
    b2 = ws.cell(row=tot_row, column=2)
    b2.fill = PatternFill("solid", fgColor=C_TOT); b2.border = BDR
    for col in range(3, ncol + 1):
        cell = ws.cell(row=tot_row, column=col, value=tot[col - 1])
        cell.font = Font(bold=True, size=9, color=("C00000" if col == ncol else C_TITLE))
        cell.fill = PatternFill("solid", fgColor=C_TOT)
        cell.number_format = FMT_NUM; cell.alignment = RIGHT; cell.border = BDR

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    for col in range(3, ncol + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "C4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _generar_excel_detalle(resultados, params):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DETALLE"

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT  = Alignment(horizontal="right",  vertical="center")
    thin   = Side(style="thin", color="BFBFBF")
    BDR    = Border(left=thin, right=thin, top=thin, bottom=thin)
    C_TITLE = "1F4E79"; C_REC = "2c6a9e"; C_EXT = "3a7abf"; C_VAL = "155724"
    C_ALT = "f7fafe"; C_TOT = "c6efce"

    def hdr(fg, sz=9):
        return Font(bold=True, color="FFFFFF", size=sz), PatternFill("solid", fgColor=fg)

    mes_nombre = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
                  "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"][int(params["mes"])]

    # Fase 2.1: el detalle ahora incluye el valor monetario por concepto.
    cols = [
        ("EMPLEADO", C_TITLE), ("DÍA", C_TITLE), ("FECHA", C_TITLE), ("D.SEM", C_TITLE),
        ("TURNO", C_TITLE), ("CICLO", C_TITLE), ("ACUM.\nINICIO", C_TITLE),
        ("TOT REC", C_REC), ("DIURNO", C_REC), ("NOCT", C_REC), ("FEST D", C_REC), ("FEST N", C_REC),
        ("TOT EXT", C_EXT), ("DIURNA", C_EXT), ("NOCT", C_EXT), ("FEST D", C_EXT), ("FEST N", C_EXT),
        ("TOTAL\nHRS", C_TITLE),
        ("$ REC DIURNO", C_VAL), ("$ REC NOCT", C_VAL), ("$ REC FEST D", C_VAL), ("$ REC FEST N", C_VAL),
        ("$ EXT DIURNA", C_VAL), ("$ EXT NOCT", C_VAL), ("$ EXT FEST D", C_VAL), ("$ EXT FEST N", C_VAL),
        ("VR. RECARGO", C_VAL), ("VR. EXTRA", C_VAL), ("VR. TOTAL", C_VAL),
    ]
    ncol = len(cols)
    last_col = get_column_letter(ncol)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"DETALLE DE NÓMINA POR TURNO/DÍA – {mes_nombre} {params['anio']} · PERSONAL DE SEGURIDAD"
    f, fill = hdr(C_TITLE, sz=12); ws["A1"].font = f; ws["A1"].fill = fill; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 24

    ws.row_dimensions[2].height = 28
    for c, (lbl, fg) in enumerate(cols, 1):
        cell = ws.cell(row=2, column=c, value=lbl)
        f, fill = hdr(fg); cell.font = f; cell.fill = fill; cell.alignment = CENTER; cell.border = BDR

    CICLO_LABEL = {"ant": "Anterior", "act": "Actual", "pos": "Posterior"}
    FMT_NUM = "#,##0"; FMT_HRS = "0.00"
    FILL_ALT = PatternFill("solid", fgColor=C_ALT)
    tot_keys = ["total_rec", "rec_diurno", "rec_nocturno", "rec_fest_d", "rec_fest_n",
                "total_ext", "ext_diurna", "ext_nocturna", "ext_fest_d", "ext_fest_n", "total_hrs",
                "vr_rec_diurno", "vr_rec_nocturno", "vr_rec_fest_d", "vr_rec_fest_n",
                "vr_ext_diurna", "vr_ext_nocturna", "vr_ext_fest_d", "vr_ext_fest_n",
                "valor_recargo", "valor_extra", "valor_total"]
    tot = {k: 0 for k in tot_keys}

    row = 3
    for r in resultados:
        for d in r.get("detalle_dias", []):
            vals = [
                r.get("nombre", ""), d.get("dia"), _fecha_dmy(d.get("fecha")), d.get("dia_semana"),
                d.get("turno"), CICLO_LABEL.get(d.get("ciclo"), d.get("ciclo")), d.get("acum_ini", 0),
                d.get("total_rec", 0), d.get("rec_diurno", 0), d.get("rec_nocturno", 0), d.get("rec_fest_d", 0), d.get("rec_fest_n", 0),
                d.get("total_ext", 0), d.get("ext_diurna", 0), d.get("ext_nocturna", 0), d.get("ext_fest_d", 0), d.get("ext_fest_n", 0),
                d.get("total_hrs", 0),
                d.get("vr_rec_diurno", 0), d.get("vr_rec_nocturno", 0), d.get("vr_rec_fest_d", 0), d.get("vr_rec_fest_n", 0),
                d.get("vr_ext_diurna", 0), d.get("vr_ext_nocturna", 0), d.get("vr_ext_fest_d", 0), d.get("vr_ext_fest_n", 0),
                d.get("valor_recargo", 0), d.get("valor_extra", 0), d.get("valor_total", 0),
            ]
            even = (row % 2 == 0)
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = BDR
                if even:
                    cell.fill = FILL_ALT
                if col == 1:
                    cell.alignment = Alignment(vertical="center")
                elif col in (2, 3, 4, 5, 6):
                    cell.alignment = CENTER
                elif col == 7 or (8 <= col <= 18):
                    cell.alignment = RIGHT; cell.number_format = FMT_HRS
                else:
                    cell.alignment = RIGHT; cell.number_format = FMT_NUM
            for k in tot_keys:
                tot[k] += d.get(k, 0)
            row += 1

    n_turnos = sum(len(r.get("detalle_dias", [])) for r in resultados)
    ws.merge_cells(f"A{row}:G{row}")
    c0 = ws.cell(row=row, column=1, value=f"TOTALES — {n_turnos} turnos")
    c0.font = Font(bold=True, size=9); c0.fill = PatternFill("solid", fgColor=C_TOT)
    c0.alignment = CENTER; c0.border = BDR
    for c in range(2, 8):
        cc = ws.cell(row=row, column=c); cc.fill = PatternFill("solid", fgColor=C_TOT); cc.border = BDR
    order = ["total_rec", "rec_diurno", "rec_nocturno", "rec_fest_d", "rec_fest_n",
             "total_ext", "ext_diurna", "ext_nocturna", "ext_fest_d", "ext_fest_n", "total_hrs",
             "vr_rec_diurno", "vr_rec_nocturno", "vr_rec_fest_d", "vr_rec_fest_n",
             "vr_ext_diurna", "vr_ext_nocturna", "vr_ext_fest_d", "vr_ext_fest_n",
             "valor_recargo", "valor_extra", "valor_total"]
    for i, k in enumerate(order):
        col = 8 + i
        es_hrs = (col <= 18)
        cell = ws.cell(row=row, column=col, value=round(tot[k], 2) if es_hrs else round(tot[k]))
        cell.font = Font(bold=True, size=9); cell.fill = PatternFill("solid", fgColor=C_TOT)
        cell.alignment = RIGHT; cell.number_format = (FMT_HRS if es_hrs else FMT_NUM); cell.border = BDR

    widths = [26, 5, 12, 7, 7, 10, 9, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 9,
              12, 12, 12, 12, 12, 12, 12, 12, 13, 13, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


_MESES_NOMBRE = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
                 "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]


def _slug_nombre(nombre):
    """Devuelve el nombre normalizado para usar como segmento de archivo."""
    if not nombre:
        return "EMPLEADO"
    norm = unicodedata.normalize("NFD", str(nombre))
    norm = norm.encode("ascii", "ignore").decode("ascii")
    norm = re.sub(r"[^A-Za-z0-9\s_-]", "", norm).strip()
    norm = re.sub(r"[\s-]+", "_", norm)
    return norm.upper() or "EMPLEADO"


def _generar_excel_soporte_empleado(r, params, reglas):
    """Fase 2.2: genera un archivo Excel de soporte individual para un empleado.

    El archivo contiene tres hojas (RESUMEN, DETALLE TURNO, CONCEPTOS) y queda
    protegido contra edición accidental.
    """
    reglas = reglas or {}
    wb = openpyxl.Workbook()

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center")
    RIGHT  = Alignment(horizontal="right",  vertical="center")
    thin   = Side(style="thin", color="BFBFBF")
    BDR    = Border(left=thin, right=thin, top=thin, bottom=thin)
    C_TITLE = "1F4E79"; C_REC = "2c6a9e"; C_EXT = "3a7abf"; C_VAL = "155724"
    C_TOT = "c6efce"; C_ALT = "f7fafe"

    def hdr(fg, sz=10):
        return Font(bold=True, color="FFFFFF", size=sz), PatternFill("solid", fgColor=fg)

    def pc(k):
        return round(float(reglas.get(k, 0)) * 100)

    FMT_NUM = "#,##0"; FMT_HRS = "0.00"
    mes = int(params["mes"]); anio = int(params["anio"])
    mes_nombre = _MESES_NOMBRE[mes]
    horas_mes = float(params.get("horas_mes", 220))
    ivh = (float(r.get("salario_mensual", 0)) / horas_mes) if horas_mes else 0

    # ── Hoja 1: RESUMEN ─────────────────────────────────────────
    ws = wb.active
    ws.title = "RESUMEN"

    ws.merge_cells("A1:D1")
    ws["A1"] = f"LIQUIDACIÓN DE NÓMINA – {mes_nombre} {anio}"
    f, fill = hdr(C_TITLE, sz=13); ws["A1"].font = f; ws["A1"].fill = fill; ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Empleado: {r.get('nombre','')} (ID {r.get('id','')})"
    ws["A2"].font = Font(bold=True, size=11); ws["A2"].alignment = LEFT

    rows_info = [
        ("Salario mensual",        round(r.get("salario_mensual", 0))),
        ("Salario proporcional",   round(r.get("salario_proporcional", 0))),
        ("Días trabajados",        r.get("dias_trabajados", 0)),
        ("Días con paga día",      r.get("dias_paga_dia", 0)),
        ("Días con aux. transp.",  r.get("dias_aux_transp", 0)),
        ("IVH (Salario ÷ horas_mes)", round(ivh)),
        ("Horas objetivo del ciclo", r.get("horas_objetivo", params.get("horas_objetivo", 132))),
        ("Inicio del ciclo (FIC)", r.get("fic", "")),
        ("Fin del ciclo (FIC_FIN)", r.get("fic_fin", "")),
        ("Inicio del siguiente ciclo", r.get("fic_siguiente", "")),
        ("", ""),
        ("Total horas liquidadas", round(r.get("total_horas", 0), 2)),
        ("Valor recargos",         r.get("valor_recargo", 0)),
        ("Valor horas extra",      r.get("valor_extra", 0)),
        ("Auxilio transporte",     r.get("auxilio_transporte", 0)),
        ("TOTAL A PAGAR",          r.get("total_pagar", 0)),
    ]
    rr = 4
    for label, value in rows_info:
        ws.cell(row=rr, column=1, value=label).font = Font(bold=(label == "TOTAL A PAGAR"), size=10)
        c = ws.cell(row=rr, column=2, value=value)
        if isinstance(value, (int, float)):
            c.number_format = (FMT_HRS if isinstance(value, float) and label.lower().startswith("horas") else FMT_NUM)
            c.alignment = RIGHT
        if label == "TOTAL A PAGAR":
            c.font = Font(bold=True, color="C00000", size=11)
            c.fill = PatternFill("solid", fgColor=C_TOT)
            ws.cell(row=rr, column=1).fill = PatternFill("solid", fgColor=C_TOT)
        rr += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 4

    # ── Hoja 2: DETALLE TURNO ───────────────────────────────────
    ws2 = wb.create_sheet("DETALLE_TURNO")

    # Soporte por empleado: sin subtotales VR. RECARGO / VR. EXTRA (causaban confusión).
    cols = [
        ("DÍA", C_TITLE), ("FECHA", C_TITLE), ("D.SEM", C_TITLE),
        ("TURNO", C_TITLE), ("CICLO", C_TITLE), ("ACUM.\nINICIO", C_TITLE),
        ("TOT REC", C_REC), ("DIURNO", C_REC), ("NOCT", C_REC), ("FEST D", C_REC), ("FEST N", C_REC),
        ("TOT EXT", C_EXT), ("DIURNA", C_EXT), ("NOCT", C_EXT), ("FEST D", C_EXT), ("FEST N", C_EXT),
        ("TOTAL\nHRS", C_TITLE),
        ("$ REC DIURNO", C_VAL), ("$ REC NOCT", C_VAL), ("$ REC FEST D", C_VAL), ("$ REC FEST N", C_VAL),
        ("$ EXT DIURNA", C_VAL), ("$ EXT NOCT", C_VAL), ("$ EXT FEST D", C_VAL), ("$ EXT FEST N", C_VAL),
        ("VR. TOTAL", C_VAL),
    ]
    ncol = len(cols)
    last_col = get_column_letter(ncol)

    ws2.merge_cells(f"A1:{last_col}1")
    ws2["A1"] = f"DETALLE TURNO – {r.get('nombre','')} – {mes_nombre} {anio}"
    f, fill = hdr(C_TITLE, sz=12); ws2["A1"].font = f; ws2["A1"].fill = fill; ws2["A1"].alignment = CENTER
    ws2.row_dimensions[1].height = 24
    ws2.row_dimensions[2].height = 28
    for c, (lbl, fg) in enumerate(cols, 1):
        cell = ws2.cell(row=2, column=c, value=lbl)
        f, fill = hdr(fg); cell.font = f; cell.fill = fill; cell.alignment = CENTER; cell.border = BDR

    CICLO_LABEL = {"ant": "Anterior", "act": "Actual", "pos": "Posterior"}
    FILL_ALT = PatternFill("solid", fgColor=C_ALT)
    tot_keys = ["total_rec", "rec_diurno", "rec_nocturno", "rec_fest_d", "rec_fest_n",
                "total_ext", "ext_diurna", "ext_nocturna", "ext_fest_d", "ext_fest_n", "total_hrs",
                "vr_rec_diurno", "vr_rec_nocturno", "vr_rec_fest_d", "vr_rec_fest_n",
                "vr_ext_diurna", "vr_ext_nocturna", "vr_ext_fest_d", "vr_ext_fest_n",
                "valor_total"]
    tot = {k: 0 for k in tot_keys}

    rr = 3
    for d in r.get("detalle_dias", []):
        vals = [
            d.get("dia"), _fecha_dmy(d.get("fecha")), d.get("dia_semana"),
            d.get("turno"), CICLO_LABEL.get(d.get("ciclo"), d.get("ciclo")), d.get("acum_ini", 0),
            d.get("total_rec", 0), d.get("rec_diurno", 0), d.get("rec_nocturno", 0), d.get("rec_fest_d", 0), d.get("rec_fest_n", 0),
            d.get("total_ext", 0), d.get("ext_diurna", 0), d.get("ext_nocturna", 0), d.get("ext_fest_d", 0), d.get("ext_fest_n", 0),
            d.get("total_hrs", 0),
            d.get("vr_rec_diurno", 0), d.get("vr_rec_nocturno", 0), d.get("vr_rec_fest_d", 0), d.get("vr_rec_fest_n", 0),
            d.get("vr_ext_diurna", 0), d.get("vr_ext_nocturna", 0), d.get("vr_ext_fest_d", 0), d.get("vr_ext_fest_n", 0),
            d.get("valor_total", 0),
        ]
        even = (rr % 2 == 0)
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=rr, column=col, value=val)
            cell.border = BDR
            if even:
                cell.fill = FILL_ALT
            if col in (1, 2, 3, 4, 5):
                cell.alignment = CENTER
            elif col == 6 or (7 <= col <= 17):
                cell.alignment = RIGHT; cell.number_format = FMT_HRS
            else:
                cell.alignment = RIGHT; cell.number_format = FMT_NUM
        for k in tot_keys:
            tot[k] += d.get(k, 0)
        rr += 1

    # Totales
    ws2.merge_cells(f"A{rr}:F{rr}")
    cT = ws2.cell(row=rr, column=1, value=f"TOTALES — {len(r.get('detalle_dias', []))} turnos")
    cT.font = Font(bold=True, size=9); cT.fill = PatternFill("solid", fgColor=C_TOT)
    cT.alignment = CENTER; cT.border = BDR
    for c in range(2, 7):
        cc = ws2.cell(row=rr, column=c); cc.fill = PatternFill("solid", fgColor=C_TOT); cc.border = BDR
    order = ["total_rec", "rec_diurno", "rec_nocturno", "rec_fest_d", "rec_fest_n",
             "total_ext", "ext_diurna", "ext_nocturna", "ext_fest_d", "ext_fest_n", "total_hrs",
             "vr_rec_diurno", "vr_rec_nocturno", "vr_rec_fest_d", "vr_rec_fest_n",
             "vr_ext_diurna", "vr_ext_nocturna", "vr_ext_fest_d", "vr_ext_fest_n",
             "valor_total"]
    for i, k in enumerate(order):
        col = 7 + i
        es_hrs = (col <= 17)
        cell = ws2.cell(row=rr, column=col, value=round(tot[k], 2) if es_hrs else round(tot[k]))
        cell.font = Font(bold=True, size=9); cell.fill = PatternFill("solid", fgColor=C_TOT)
        cell.alignment = RIGHT; cell.number_format = (FMT_HRS if es_hrs else FMT_NUM); cell.border = BDR

    widths = [5, 12, 7, 7, 10, 9, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 9,
              12, 12, 12, 12, 12, 12, 12, 12, 14]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A3"

    # ── Hoja 3: CONCEPTOS (horas + valor por concepto) ─────────
    ws3 = wb.create_sheet("CONCEPTOS")
    ws3.merge_cells("A1:D1")
    ws3["A1"] = f"HORAS Y VALOR POR CONCEPTO – {r.get('nombre','')} – {mes_nombre} {anio}"
    f, fill = hdr(C_TITLE, sz=12); ws3["A1"].font = f; ws3["A1"].fill = fill; ws3["A1"].alignment = CENTER

    headers = ["CONCEPTO", "FACTOR", "HORAS", "VALOR"]
    for c, h in enumerate(headers, 1):
        cell = ws3.cell(row=2, column=c, value=h)
        f, fill = hdr(C_TITLE); cell.font = f; cell.fill = fill; cell.alignment = CENTER; cell.border = BDR

    # Conceptos sin % en el nombre — el porcentaje lo expresa la columna FACTOR,
    # que se lee dinámicamente de la hoja de Parámetros (reglas) en cada generación.
    val = r.get("val", {}) or {}
    filas_conc = [
        ("Recargo diurno",            "REC_DIURNO",            0.0,
         r.get("hrs_diurnas", 0),     val.get("rec_diurno", 0)),
        ("Recargo nocturno",          "REC_NOCTURNO",          0.35,
         r.get("hrs_nocturnas", 0),   val.get("rec_noct", 0)),
        ("Recargo dom/fest diurno",   "REC_DOM_FEST_DIURNO",   0.80,
         r.get("hrs_fest_diurnas", 0), val.get("rec_fest_d", 0)),
        ("Recargo dom/fest nocturno", "REC_DOM_FEST_NOCTURNO", 1.15,
         r.get("hrs_fest_noc", 0),    val.get("rec_fest_n", 0)),
        ("Extra diurna",              "EXT_DIURNA",            1.25,
         r.get("hrs_ext_diurnas", 0), val.get("ext_diurna", 0)),
        ("Extra nocturna",            "EXT_NOCTURNA",          1.75,
         r.get("hrs_ext_noc", 0),     val.get("ext_noct", 0)),
        ("Extra fest diurna",         "EXT_FEST_DIURNA",       2.05,
         0,                           val.get("ext_fest_d", 0)),
        ("Extra fest nocturna",       "EXT_FEST_NOCTURNA",     2.55,
         0,                           val.get("ext_fest_n", 0)),
    ]
    rr = 3
    for concepto, regla_key, default, horas, vlr in filas_conc:
        factor = reglas.get(regla_key, default)
        ws3.cell(row=rr, column=1, value=concepto).alignment = LEFT
        ws3.cell(row=rr, column=2, value=f"{round(factor*100)}%").alignment = CENTER
        cH = ws3.cell(row=rr, column=3, value=horas);  cH.number_format = FMT_HRS; cH.alignment = RIGHT
        cV = ws3.cell(row=rr, column=4, value=vlr);    cV.number_format = FMT_NUM; cV.alignment = RIGHT
        for c in range(1, 5):
            ws3.cell(row=rr, column=c).border = BDR
        rr += 1
    # Totales
    ws3.cell(row=rr, column=1, value="TOTAL").font = Font(bold=True)
    ws3.cell(row=rr, column=1).fill = PatternFill("solid", fgColor=C_TOT)
    cV = ws3.cell(row=rr, column=4, value=r.get("valor_recargo", 0) + r.get("valor_extra", 0))
    cV.font = Font(bold=True, color="C00000"); cV.number_format = FMT_NUM; cV.alignment = RIGHT
    cV.fill = PatternFill("solid", fgColor=C_TOT)
    for c in range(2, 4):
        cell = ws3.cell(row=rr, column=c); cell.fill = PatternFill("solid", fgColor=C_TOT); cell.border = BDR
    ws3.cell(row=rr, column=1).border = BDR
    cV.border = BDR

    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 16

    # Proteger las hojas contra edición accidental (sin contraseña).
    for sheet in (ws, ws2, ws3):
        sheet.protection.sheet = True
        sheet.protection.enable()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _generar_zip_soportes(resultados, params, reglas):
    """Empaqueta los soportes individuales en un ZIP en memoria."""
    zip_buf = io.BytesIO()
    nombres_usados = {}
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in resultados:
            slug = _slug_nombre(r.get("nombre", "")) or f"EMPLEADO_{r.get('id','')}"
            # Evitar colisiones por nombres repetidos.
            count = nombres_usados.get(slug, 0) + 1
            nombres_usados[slug] = count
            suffix = "" if count == 1 else f"_{count}"
            filename = f"LIQUIDACION_{slug}{suffix}.xlsx"
            buf = _generar_excel_soporte_empleado(r, params, reglas)
            zf.writestr(filename, buf.getvalue())
    zip_buf.seek(0)
    return zip_buf


# ─────────────────────────────────────────────────────────────
# Rutas
# ─────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/defaults")
def api_defaults():
    return jsonify(_load_current())


@app.route("/api/guardar-params", methods=["POST"])
def api_guardar_params():
    try:
        body = request.get_json(force=True) or {}
        cfg = {
            "params": body.get("params", {}),
            "reglas": body.get("reglas", {}),
            "festivos": body.get("festivos", []),
        }
        _save_config(cfg)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/api/calcular", methods=["POST"])
def api_calcular():
    body = request.get_json(force=True)
    try:
        empleados_limpios, validation = _validar_turnos_empleados(body["empleados"])
        if validation["errors"]:
            return jsonify({"ok": False, "error": "Hay turnos inválidos en la plantilla.", "validation": validation}), 400
        resultados = calcular_nomina(
            params=body["params"],
            empleados=empleados_limpios,
            reglas=body["reglas"],
            festivos=body["festivos"],
        )
        return jsonify({"ok": True, "resultados": resultados, "validation": validation})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/api/descargar", methods=["POST"])
def api_descargar():
    body = request.get_json(force=True)
    try:
        empleados_limpios, validation = _validar_turnos_empleados(body["empleados"])
        if validation["errors"]:
            return jsonify({"ok": False, "error": "Hay turnos inválidos en la plantilla.", "validation": validation}), 400
        resultados = calcular_nomina(
            params=body["params"],
            empleados=empleados_limpios,
            reglas=body["reglas"],
            festivos=body["festivos"],
        )
        buf = _generar_excel(resultados, body["params"], body.get("reglas", {}))
        mes = int(body["params"]["mes"])
        anio = int(body["params"]["anio"])
        filename = f"Liquidacion_Nomina_{anio}_{mes:02d}.xlsx"
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/api/descargar-soportes", methods=["POST"])
def api_descargar_soportes():
    """Fase 2.2: genera un Excel individual por empleado y los empaqueta en ZIP.

    Si solo hay un empleado, devuelve el XLSX directamente.
    """
    body = request.get_json(force=True)
    try:
        empleados_limpios, validation = _validar_turnos_empleados(body["empleados"])
        if validation["errors"]:
            return jsonify({"ok": False, "error": "Hay turnos inválidos en la plantilla.", "validation": validation}), 400
        resultados = calcular_nomina(
            params=body["params"],
            empleados=empleados_limpios,
            reglas=body["reglas"],
            festivos=body["festivos"],
        )
        if not resultados:
            return jsonify({"ok": False, "error": "No hay empleados para generar soportes."}), 400
        params = body["params"]
        reglas = body.get("reglas", {})
        mes = int(params["mes"]); anio = int(params["anio"])
        if len(resultados) == 1:
            r = resultados[0]
            slug = _slug_nombre(r.get("nombre", "")) or f"EMPLEADO_{r.get('id','')}"
            filename = f"LIQUIDACION_{slug}.xlsx"
            buf = _generar_excel_soporte_empleado(r, params, reglas)
            return send_file(
                buf,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename,
            )
        zip_buf = _generar_zip_soportes(resultados, params, reglas)
        mes_nombre = _MESES_NOMBRE[mes]
        zip_filename = f"Soportes_Nomina_{mes_nombre}_{anio}.zip"
        return send_file(
            zip_buf,
            mimetype="application/zip",
            as_attachment=True,
            download_name=zip_filename,
        )
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/api/auditoria", methods=["POST"])
def api_auditoria():
    """Fase 2.4: auditoría de turnos con múltiples conceptos y cruces críticos."""
    body = request.get_json(force=True)
    try:
        empleados_limpios, validation = _validar_turnos_empleados(body["empleados"])
        if validation["errors"]:
            return jsonify({"ok": False, "error": "Hay turnos inválidos en la plantilla.", "validation": validation}), 400
        resultados = calcular_nomina(
            params=body["params"],
            empleados=empleados_limpios,
            reglas=body["reglas"],
            festivos=body["festivos"],
        )
        params = body["params"]
        festivos_set = set((body.get("festivos") or []))
        anio = int(params["anio"]); mes = int(params["mes"])
        umbral_h = float(params.get("horas_objetivo", 132))

        hallazgos = []
        kpi = {
            "turnos_revisados": 0,
            "turnos_multi_concepto": 0,
            "turnos_cruce_dia": 0,
            "turnos_cruce_ciclo": 0,
            "turnos_cruce_mes": 0,
            "turnos_post_umbral": 0,
            "turnos_dom_fest": 0,
        }
        for r in resultados:
            for d in r.get("detalle_dias", []):
                kpi["turnos_revisados"] += 1
                conceptos = 0
                detalles = []
                # Inspecciona cuántos conceptos no-nulos contiene el turno (>1 = multi-concepto).
                for k, lbl in (("rec_diurno","Rec.Diurno"), ("rec_nocturno","Rec.Noct"),
                               ("rec_fest_d","Rec.FestD"), ("rec_fest_n","Rec.FestN"),
                               ("ext_diurna","Ext.Diurna"), ("ext_nocturna","Ext.Noct"),
                               ("ext_fest_d","Ext.FestD"), ("ext_fest_n","Ext.FestN")):
                    val = d.get(k, 0) or 0
                    if val > 0:
                        conceptos += 1
                        detalles.append(f"{lbl}={round(val,2)}h")
                if conceptos > 1:
                    kpi["turnos_multi_concepto"] += 1

                # ¿Es un sábado/dom/fest?
                fecha_str = d.get("fecha", "")
                try:
                    fecha_obj = datetime.strptime(fecha_str[:10], "%Y-%m-%d").date()
                except Exception:
                    fecha_obj = None
                es_dom_fest = False
                if fecha_obj:
                    es_dom_fest = (fecha_obj.weekday() == 6) or (fecha_str[:10] in festivos_set)
                if es_dom_fest:
                    kpi["turnos_dom_fest"] += 1

                # Detecta superación del umbral del ciclo:
                acum_ini = float(d.get("acum_ini", 0) or 0)
                total_rec = float(d.get("total_rec", 0) or 0)
                total_ext = float(d.get("total_ext", 0) or 0)
                if acum_ini >= umbral_h and total_rec > 0:
                    kpi["turnos_post_umbral"] += 1
                    hallazgos.append({
                        "tipo": "post_umbral",
                        "severidad": "alta",
                        "empleado": r.get("nombre"),
                        "id": r.get("id"),
                        "dia": d.get("dia"),
                        "fecha": d.get("fecha"),
                        "ciclo": d.get("ciclo"),
                        "mensaje": (f"Turno con {acum_ini}h acumuladas (≥{umbral_h}) liquidó "
                                    f"{total_rec}h como recargo en vez de extra."),
                    })

                # Identifica cruces de mes / ciclo / día por turnos N (clasificación heurística por código).
                turno_cod = (d.get("turno") or "").upper()
                if fecha_obj and turno_cod in {"N", "FN", "AN"}:
                    kpi["turnos_cruce_dia"] += 1
                    siguiente = fecha_obj.replace(day=1) if False else fecha_obj
                    # cruce de mes
                    if fecha_obj.day == (28 if mes == 2 and anio % 4 else 30 if mes in {4,6,9,11} else 31):
                        kpi["turnos_cruce_mes"] += 1
                        hallazgos.append({
                            "tipo": "cruce_mes",
                            "severidad": "info",
                            "empleado": r.get("nombre"),
                            "id": r.get("id"),
                            "dia": d.get("dia"),
                            "fecha": d.get("fecha"),
                            "mensaje": "Turno nocturno que cruza al siguiente mes — verificar prorrateo.",
                        })

                # cruce de ciclo: turno cuya fecha == fic_fin del empleado
                if d.get("fecha") and r.get("fic_fin") == d.get("fecha"):
                    kpi["turnos_cruce_ciclo"] += 1
                    if conceptos > 1:
                        hallazgos.append({
                            "tipo": "cruce_ciclo",
                            "severidad": "info",
                            "empleado": r.get("nombre"),
                            "id": r.get("id"),
                            "dia": d.get("dia"),
                            "fecha": d.get("fecha"),
                            "mensaje": (f"Turno en el último día del ciclo combina {conceptos} conceptos: "
                                        + ", ".join(detalles) + "."),
                        })

                # multi-concepto que también es sábado/dom/fest — incluir en hallazgos informativos
                if conceptos > 1 and es_dom_fest:
                    hallazgos.append({
                        "tipo": "multi_concepto",
                        "severidad": "info",
                        "empleado": r.get("nombre"),
                        "id": r.get("id"),
                        "dia": d.get("dia"),
                        "fecha": d.get("fecha"),
                        "mensaje": (f"Turno en domingo/festivo con {conceptos} conceptos: "
                                    + ", ".join(detalles) + "."),
                    })

        # Festivos fuera del mes liquidado (debilidad detectada en Fase 1).
        festivos_fuera = []
        for f in (body.get("festivos") or []):
            try:
                fd = datetime.strptime(f, "%Y-%m-%d").date()
            except Exception:
                continue
            if fd.year != anio or fd.month != mes:
                festivos_fuera.append(f)

        return jsonify({
            "ok": True,
            "kpi": kpi,
            "hallazgos": hallazgos,
            "festivos_fuera_periodo": festivos_fuera,
            "umbral_horas": umbral_h,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/api/descargar-detalle", methods=["POST"])
def api_descargar_detalle():
    body = request.get_json(force=True)
    try:
        empleados_limpios, validation = _validar_turnos_empleados(body["empleados"])
        if validation["errors"]:
            return jsonify({"ok": False, "error": "Hay turnos inválidos en la plantilla.", "validation": validation}), 400
        resultados = calcular_nomina(
            params=body["params"],
            empleados=empleados_limpios,
            reglas=body["reglas"],
            festivos=body["festivos"],
        )
        buf = _generar_excel_detalle(resultados, body["params"])
        mes = int(body["params"]["mes"])
        anio = int(body["params"]["anio"])
        filename = f"Detalle_Nomina_{anio}_{mes:02d}.xlsx"
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/api/plantilla")
def api_plantilla():
    fp = BASE_DIR / "Plantilla_Nomina.xlsx"
    if not fp.exists():
        return jsonify({"ok": False, "error": "Plantilla no encontrada"}), 404
    return send_file(str(fp), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="Plantilla_Nomina.xlsx")


@app.route("/api/cargar-excel", methods=["POST"])
def api_cargar_excel():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No se recibió archivo"}), 400
    f = request.files["file"]
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        data = _excel_to_data(wb)
        return jsonify({"ok": True, "data": data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


# Inicializar DB al importar (gunicorn no llama __main__)
_init_db()

if __name__ == "__main__":
    print("\n✅  Servidor de nómina arrancado")
    print("   Abre tu navegador en: http://localhost:5050\n")
    app.run(host="0.0.0.0", port=5050, debug=False)
